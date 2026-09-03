"""Contrat d'export vers le gestionnaire externe (ch. 13).

Ces vues sont le véritable point de livraison du projet. Leur structure doit
rester figée une fois validée : les tests ci-dessous verrouillent à la fois
le périmètre (données validées uniquement) et la présence des colonnes.
"""

import io

import pytest
from openpyxl import load_workbook

from tests.conftest import horodatage, nouvel_id


def _creer_rotation(client, entetes, parc, *, validee: bool) -> str:
    identifiant = nouvel_id()
    client.post(
        "/api/v1/marinage/rotations",
        json={
            "id": identifiant,
            "dumper_id": str(parc["engins"]["DU01"].id),
            "site_id": parc["site_id"],
            "horodatage": horodatage(1),
            "quantite_estimee_t": 28.5,
            "nature_quantite": "estimation",
        },
        headers=entetes("agent"),
    )
    if validee:
        for statut in ("controlee", "validee"):
            client.post(
                f"/api/v1/validation/rotation_dumper/{identifiant}/statut",
                json={"nouveau_statut": statut},
                headers=entetes("controleur"),
            )
    return identifiant


def test_seules_les_donnees_validees_sont_exposees(client, entetes, parc):
    """Une donnée brute ne doit jamais atteindre le gestionnaire externe."""
    validee = _creer_rotation(client, entetes, parc, validee=True)
    brute = _creer_rotation(client, entetes, parc, validee=False)

    export = client.get("/api/v1/exports/rotations", headers=entetes("controleur")).json()
    identifiants = {ligne["id"] for ligne in export["lignes"]}

    assert validee in identifiants
    assert brute not in identifiants


def test_le_catalogue_declare_les_six_vues_du_contrat(client, entetes):
    catalogue = client.get(
        "/api/v1/exports/catalogue", headers=entetes("controleur")
    ).json()
    noms = {e["nom"] for e in catalogue["exports"]}
    assert noms == {
        "foration",
        "activite_engin",
        "rotations",
        "pesees",
        "charges_engin",
        "completude",
    }


@pytest.mark.parametrize(
    ("export", "colonnes_attendues"),
    [
        (
            "foration",
            {
                "id_trou",
                "site",
                "numero_tir",
                "foreuse",
                "operateur",
                "date_foration",
                "duree_heures",
                "utilisation_foreuse",
                "metres_lineaires",
                "numero_taillant",
                "statut",
            },
        ),
        (
            "activite_engin",
            {
                "engin",
                "famille",
                "site",
                "centre_cout_reel",
                "type_evenement",
                "horodatage",
                "compteur",
                "cause_code",
                "carburant_litres",
                "statut",
            },
        ),
        (
            "rotations",
            {
                "dumper",
                "site",
                "horodatage",
                "point_deversement",
                "poids_reel_t",
                "quantite_estimee_t",
                "nature_quantite",
                "passage_precedent",
                "statut",
            },
        ),
        (
            "pesees",
            {"site", "horodatage", "client", "immatriculation", "produit", "poids_t", "numero_bon"},
        ),
        (
            "charges_engin",
            {
                "engin",
                "famille",
                "site",
                "nature",
                "categorie",
                "date_charge",
                "montant",
                "periode_debut",
                "periode_fin",
            },
        ),
        (
            "completude",
            {"site", "jour", "trous_declares", "trous_valides", "trous_non_clotures"},
        ),
    ],
)
def test_les_colonnes_du_contrat_sont_presentes(client, entetes, export, colonnes_attendues):
    """Le gestionnaire valide ces colonnes une fois ; elles ne doivent plus bouger.

    Ce test échoue si une colonne du contrat disparaît ou est renommée — ce
    qui est exactement l'alerte attendue.
    """
    reponse = client.get(f"/api/v1/exports/{export}", headers=entetes("controleur"))
    assert reponse.status_code == 200
    colonnes = set(reponse.json()["colonnes"])
    manquantes = colonnes_attendues - colonnes
    assert not manquantes, f"Colonnes disparues de l'export {export} : {sorted(manquantes)}"


def test_l_export_rotations_n_expose_aucun_tonnage_total(client, entetes, parc):
    """Aucune colonne ne doit additionner un tonnage pesé et un tonnage estimé."""
    colonnes = client.get(
        "/api/v1/exports/rotations", headers=entetes("controleur")
    ).json()["colonnes"]
    suspectes = [c for c in colonnes if "total" in c.lower()]
    assert not suspectes, f"Colonnes agrégeant réel et estimé : {suspectes}"


def test_le_filtre_par_site_et_par_periode(client, entetes, parc):
    from datetime import UTC, datetime, timedelta

    _creer_rotation(client, entetes, parc, validee=True)
    h = entetes("controleur")
    aujourdhui = datetime.now(UTC).date()

    sur_site = client.get(
        "/api/v1/exports/rotations", params={"site": "KOS"}, headers=h
    ).json()
    assert sur_site["nb_lignes"] == 1

    autre_site = client.get(
        "/api/v1/exports/rotations", params={"site": "BKE"}, headers=h
    ).json()
    assert autre_site["nb_lignes"] == 0

    # La borne haute doit être inclusive : « au 31/08 » contient le 31/08 au soir.
    dans_la_periode = client.get(
        "/api/v1/exports/rotations",
        params={"du": aujourdhui.isoformat(), "au": aujourdhui.isoformat()},
        headers=h,
    ).json()
    assert dans_la_periode["nb_lignes"] == 1

    avant = client.get(
        "/api/v1/exports/rotations",
        params={"au": (aujourdhui - timedelta(days=1)).isoformat()},
        headers=h,
    ).json()
    assert avant["nb_lignes"] == 0


def test_le_fichier_excel_porte_ses_donnees_et_son_perimetre(client, entetes, parc):
    """Un export qui circule sans son périmètre finit par être mal lu."""
    _creer_rotation(client, entetes, parc, validee=True)

    reponse = client.get(
        "/api/v1/exports/rotations/fichier",
        params={"format": "xlsx", "site": "KOS"},
        headers=entetes("controleur"),
    )
    assert reponse.status_code == 200
    assert "attachment" in reponse.headers["content-disposition"]

    classeur = load_workbook(io.BytesIO(reponse.content))
    assert classeur.sheetnames == ["Donnees", "Contexte"]
    assert classeur["Donnees"].max_row == 2  # en-tête + une rotation

    contexte = {
        ligne[0]: ligne[1] for ligne in classeur["Contexte"].iter_rows(values_only=True)
    }
    assert "validée" in contexte["Périmètre"]
    assert contexte["Site"] == "KOS"


def test_le_csv_est_lisible_par_excel_en_francais(client, entetes, parc):
    """Séparateur point-virgule et BOM UTF-8 : sans eux, Excel écrase les
    accents et met toute la ligne dans une seule colonne."""
    _creer_rotation(client, entetes, parc, validee=True)

    reponse = client.get(
        "/api/v1/exports/rotations/fichier",
        params={"format": "csv"},
        headers=entetes("controleur"),
    )
    assert reponse.status_code == 200
    assert reponse.content.startswith(b"\xef\xbb\xbf")

    entete = reponse.content.decode("utf-8-sig").splitlines()[0]
    assert ";" in entete


def test_un_agent_de_terrain_ne_telecharge_pas_les_exports(client, entetes, parc):
    reponse = client.get(
        "/api/v1/exports/rotations/fichier", params={"format": "csv"}, headers=entetes("agent")
    )
    assert reponse.status_code == 403


def test_un_export_inconnu_retourne_404(client, entetes):
    reponse = client.get("/api/v1/exports/inexistant", headers=entetes("controleur"))
    assert reponse.status_code == 404


def test_l_export_pdf_est_un_pdf_valide(client, entetes, parc, session):
    """Le PDF est un document imprimable, pas un jeu de données."""
    reponse = client.get(
        "/api/v1/exports/rotations/fichier",
        params={"format": "pdf"},
        headers=entetes("controleur"),
    )
    assert reponse.status_code == 200, reponse.text
    assert reponse.headers["content-type"] == "application/pdf"
    assert reponse.content.startswith(b"%PDF-")
    assert ".pdf" in reponse.headers["content-disposition"]


def test_le_pdf_ne_retient_qu_une_partie_des_colonnes(client, entetes, parc):
    """Vingt-six colonnes ne tiennent pas sur une page : la sélection est
    déclarée dans le catalogue, pas laissée au hasard."""
    from app.exports.generateur import _colonnes_pdf
    from app.exports.vues import CATALOGUE

    for nom, definition in CATALOGUE.items():
        assert definition.colonnes_pdf, f"{nom} n'a pas de colonnes PDF déclarées"
        retenues = _colonnes_pdf(definition, list(definition.colonnes_pdf))
        assert retenues, nom
        assert len(retenues) <= 12, f"{nom} : trop de colonnes pour une page"


def test_un_format_inconnu_est_refuse(client, entetes, parc):
    reponse = client.get(
        "/api/v1/exports/rotations/fichier",
        params={"format": "docx"},
        headers=entetes("controleur"),
    )
    assert reponse.status_code == 422
