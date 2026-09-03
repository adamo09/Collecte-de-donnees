"""Production des fichiers CSV, Excel et PDF à partir des vues d'export."""

import csv
import io
import uuid
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.exports.vues import CATALOGUE, DefinitionVue

# Séparateur point-virgule : Excel en configuration française n'ouvre pas
# correctement un CSV séparé par des virgules.
SEPARATEUR_CSV = ";"


class VueInconnue(Exception):
    """La vue demandée n'appartient pas au catalogue d'export."""


def resoudre_vue(nom: str) -> DefinitionVue:
    definition = CATALOGUE.get(nom)
    if definition is None:
        raise VueInconnue(
            f"« {nom} » n'est pas une vue d'export. Vues disponibles : "
            + ", ".join(sorted(CATALOGUE))
        )
    return definition


def construire_requete(
    definition: DefinitionVue,
    site: str | None = None,
    du: date | None = None,
    au: date | None = None,
    filtres: dict[str, Any] | None = None,
    limite: int | None = None,
) -> tuple[str, dict[str, Any]]:
    """Assemble la requête filtrée.

    Les noms de colonnes proviennent exclusivement du catalogue, jamais de
    l'utilisateur : aucune valeur saisie n'est interpolée dans le SQL.
    """
    conditions: list[str] = []
    parametres: dict[str, Any] = {}

    if site and definition.colonne_site:
        conditions.append(f"{definition.colonne_site} = :site")
        parametres["site"] = site

    if definition.colonne_date:
        if du is not None:
            conditions.append(f"{definition.colonne_date} >= :du")
            parametres["du"] = du
        if au is not None:
            # Borne haute inclusive, y compris pour une colonne horodatée.
            conditions.append(f"{definition.colonne_date} < :au")
            parametres["au"] = au

    for nom_filtre, valeur in (filtres or {}).items():
        colonne = definition.filtres_supplementaires.get(nom_filtre)
        if colonne is None or valeur is None:
            continue
        conditions.append(f"{colonne} = :f_{nom_filtre}")
        parametres[f"f_{nom_filtre}"] = valeur

    clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    clause_limite = f"LIMIT {int(limite)}" if limite else ""
    requete = (
        f"SELECT * FROM {definition.vue_sql} {clause} "
        f"ORDER BY {definition.tri} {clause_limite}"
    )
    return requete, parametres


def lire(
    session: Session,
    definition: DefinitionVue,
    site: str | None = None,
    du: date | None = None,
    au: date | None = None,
    filtres: dict[str, Any] | None = None,
    limite: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    requete, parametres = construire_requete(definition, site, du, au, filtres, limite)
    resultat = session.execute(text(requete), parametres)
    colonnes = list(resultat.keys())
    lignes = [dict(ligne) for ligne in resultat.mappings()]
    return colonnes, lignes


def _valeur_texte(valeur: Any) -> str:
    """Rend une valeur au format attendu par Excel en configuration française."""
    if valeur is None:
        return ""
    if isinstance(valeur, bool):
        return "oui" if valeur else "non"
    if isinstance(valeur, datetime):
        return valeur.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(valeur, date):
        return valeur.strftime("%Y-%m-%d")
    if isinstance(valeur, timedelta):
        # Une durée est plus exploitable en heures décimales qu'en
        # « 0:38:00 », que le tableur traite comme du texte.
        return f"{valeur.total_seconds() / 3600:.4f}".replace(".", ",")
    if isinstance(valeur, Decimal | float):
        return str(valeur).replace(".", ",")
    if isinstance(valeur, Enum):
        return str(valeur.value)
    return str(valeur)


def _valeur_excel(valeur: Any) -> Any:
    """Convertit une valeur en un type qu'openpyxl sait écrire.

    Les nombres et les dates restent natifs pour rester calculables dans le
    tableur ; tout le reste — UUID, énumérations, JSON — devient du texte.
    """
    if valeur is None or isinstance(valeur, bool | int | float | str):
        return valeur
    if isinstance(valeur, Decimal):
        return float(valeur)
    if isinstance(valeur, datetime):
        # Excel ne sait pas représenter un fuseau horaire : l'horodatage est
        # normalisé en UTC, puis rendu naïf.
        return valeur.astimezone(UTC).replace(tzinfo=None) if valeur.tzinfo else valeur
    if isinstance(valeur, date):
        return valeur
    if isinstance(valeur, timedelta):
        return round(valeur.total_seconds() / 3600, 4)
    if isinstance(valeur, Enum):
        return str(valeur.value)
    if isinstance(valeur, uuid.UUID):
        return str(valeur)
    return str(valeur)


def generer_csv(colonnes: list[str], lignes: list[dict[str, Any]]) -> bytes:
    tampon = io.StringIO()
    redacteur = csv.writer(tampon, delimiter=SEPARATEUR_CSV, lineterminator="\r\n")
    redacteur.writerow(colonnes)
    for ligne in lignes:
        redacteur.writerow([_valeur_texte(ligne.get(c)) for c in colonnes])
    # BOM UTF-8 : sans lui, Excel affiche « Bouaké » en « BouakÃ© ».
    return tampon.getvalue().encode("utf-8-sig")


def generer_excel(
    definition: DefinitionVue,
    colonnes: list[str],
    lignes: list[dict[str, Any]],
    contexte: dict[str, Any] | None = None,
) -> bytes:
    """Produit un classeur avec une feuille de données et une feuille de contexte.

    La feuille de contexte n'est pas décorative : elle indique le périmètre
    exact du fichier et rappelle que seules des données validées y figurent.
    Un export qui circule sans son périmètre finit toujours par être mal lu.
    """
    classeur = Workbook()

    feuille = classeur.active
    feuille.title = "Donnees"

    # Marine de la charte CADERAC. C'est le seul endroit où la couleur
    # de la marque quitte l'application pour voyager dans un fichier.
    entete_fond = PatternFill("solid", fgColor="003559")
    entete_police = Font(color="FFFFFF", bold=True)

    feuille.append(colonnes)
    for cellule in feuille[1]:
        cellule.fill = entete_fond
        cellule.font = entete_police
        cellule.alignment = Alignment(horizontal="center", vertical="center")

    for ligne in lignes:
        feuille.append([_valeur_excel(ligne.get(c)) for c in colonnes])

    for index, nom in enumerate(colonnes, start=1):
        largeur = max(len(nom) + 2, 12)
        feuille.column_dimensions[get_column_letter(index)].width = min(largeur, 40)
    feuille.freeze_panes = "A2"
    if lignes:
        feuille.auto_filter.ref = (
            f"A1:{get_column_letter(len(colonnes))}{len(lignes) + 1}"
        )

    contexte_feuille = classeur.create_sheet("Contexte")
    infos = [
        ("Export", definition.libelle),
        ("Vue source", definition.vue_sql),
        ("Description", definition.description),
        ("Généré le", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Nombre de lignes", len(lignes)),
        (
            "Périmètre",
            "Données au statut « validée » uniquement"
            if definition.nom != "completude"
            else "Tous statuts — vue de pilotage",
        ),
    ]
    for cle, valeur in (contexte or {}).items():
        infos.append((cle, valeur))

    for cle, valeur in infos:
        contexte_feuille.append([cle, valeur])
    for cellule in contexte_feuille["A"]:
        cellule.font = Font(bold=True)
    contexte_feuille.column_dimensions["A"].width = 22
    contexte_feuille.column_dimensions["B"].width = 70

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


# =====================================================================
# Édition PDF
# =====================================================================

# Un PDF de vingt-six colonnes ne se lit pas. Le PDF n'est donc pas un
# jeu de données mais un document : on l'imprime, on le classe, on le
# joint à un dossier. Les formats de travail restent le classeur Excel
# et le CSV, qui portent l'intégralité des colonnes.
LIGNES_MAX_PDF = 1500

MARINE = colors.HexColor("#003559")
CIEL_PALE = colors.HexColor("#E4F0F7")
ENCRE = colors.HexColor("#14202E")
ENCRE_DOUX = colors.HexColor("#4A5768")
FILET = colors.HexColor("#C9D3DF")

_LOGO = Path(__file__).parent / "assets" / "logo-caderac.png"


def _colonnes_pdf(definition: DefinitionVue, colonnes: list[str]) -> list[str]:
    """Colonnes retenues pour l'édition, dans l'ordre de la vue."""
    if not definition.colonnes_pdf:
        return colonnes[:10]
    retenues = [c for c in definition.colonnes_pdf if c in colonnes]
    return retenues or colonnes[:10]


def _texte_pdf(valeur: Any) -> str:
    """Rend une valeur imprimable, sans jamais afficher « None »."""
    if valeur is None:
        return "—"
    if isinstance(valeur, datetime):
        return valeur.strftime("%d/%m/%y %H:%M")
    if isinstance(valeur, date):
        return valeur.strftime("%d/%m/%Y")
    if isinstance(valeur, Enum):
        return str(valeur.value)
    if isinstance(valeur, Decimal | float):
        return f"{valeur:,.2f}".replace(",", " ").replace(".", ",")
    return str(valeur)


def generer_pdf(
    definition: DefinitionVue,
    colonnes: list[str],
    lignes: list[dict[str, Any]],
    contexte: dict[str, Any] | None = None,
) -> bytes:
    """Produit un document paysage : en-tête, périmètre, puis le tableau.

    Le périmètre figure en tête et non en annexe : un PDF circule par
    courriel et finit par être lu sans son message d'accompagnement.
    """
    retenues = _colonnes_pdf(definition, colonnes)
    tronque = len(lignes) > LIGNES_MAX_PDF
    visibles = lignes[:LIGNES_MAX_PDF]

    tampon = io.BytesIO()
    document = SimpleDocTemplate(
        tampon,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=16 * mm,
        title=f"CADERAC — {definition.libelle}",
        author="CADERAC — Contrôle de gestion",
        subject=definition.description,
    )

    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle(
        "titre", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=17, leading=20, textColor=MARINE, alignment=0, spaceAfter=2,
    )
    style_soustitre = ParagraphStyle(
        "soustitre", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, leading=12, textColor=ENCRE_DOUX, spaceAfter=0,
    )
    style_cellule = ParagraphStyle(
        "cellule", parent=styles["Normal"], fontName="Helvetica",
        fontSize=6.6, leading=8.2, textColor=ENCRE,
    )
    style_entete = ParagraphStyle(
        "entete", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=6.4, leading=7.8, textColor=colors.white,
    )
    style_note = ParagraphStyle(
        "note", parent=styles["Normal"], fontName="Helvetica-Oblique",
        fontSize=7.5, leading=10, textColor=ENCRE_DOUX, spaceBefore=6,
    )

    histoire: list[Any] = []

    # --- Bandeau : logo à gauche, titre et périmètre à droite ---------
    infos = {
        "Généré le": datetime.now().strftime("%d/%m/%Y à %H:%M"),
        "Lignes": f"{len(lignes)}",
        "Périmètre": (
            "Données validées uniquement"
            if definition.nom != "completude"
            else "Tous statuts — vue de pilotage"
        ),
    }
    infos.update({c: _texte_pdf(v) for c, v in (contexte or {}).items()})
    lignes_infos = " · ".join(f"<b>{c}</b> {v}" for c, v in infos.items())

    entete_droite = [
        Paragraph(definition.libelle, style_titre),
        Paragraph(definition.description, style_soustitre),
        Spacer(1, 4),
        Paragraph(lignes_infos, style_soustitre),
    ]
    if _LOGO.exists():
        logo = Image(str(_LOGO), width=26 * mm, height=21 * mm, kind="proportional")
        bandeau = Table([[logo, entete_droite]], colWidths=[32 * mm, None])
    else:
        bandeau = Table([[entete_droite]], colWidths=[None])
    bandeau.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LINEBELOW", (0, 0), (-1, -1), 1.2, MARINE),
        ])
    )
    histoire.append(bandeau)
    histoire.append(Spacer(1, 8))

    # --- Le tableau ----------------------------------------------------
    if visibles:
        donnees = [[Paragraph(c.replace("_", " "), style_entete) for c in retenues]]
        for ligne in visibles:
            donnees.append(
                [Paragraph(_texte_pdf(ligne.get(c)), style_cellule) for c in retenues]
            )
        largeur = document.width / len(retenues)
        tableau = Table(donnees, colWidths=[largeur] * len(retenues), repeatRows=1)
        tableau.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), MARINE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CIEL_PALE]),
                ("GRID", (0, 0), (-1, -1), 0.25, FILET),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ])
        )
        histoire.append(tableau)
    else:
        histoire.append(
            Paragraph(
                "Aucune donnée ne correspond à ce périmètre.", style_soustitre
            )
        )

    # --- Mentions ------------------------------------------------------
    ecartees = len(colonnes) - len(retenues)
    if ecartees > 0:
        histoire.append(
            Paragraph(
                f"Édition abrégée : {len(retenues)} colonnes sur {len(colonnes)}. "
                "Le classeur Excel et le fichier CSV portent l’intégralité des "
                "colonnes, y compris les métadonnées de traçabilité.",
                style_note,
            )
        )
    if tronque:
        histoire.append(
            Paragraph(
                f"Document tronqué aux {LIGNES_MAX_PDF} premières lignes sur "
                f"{len(lignes)}. Télécharger le classeur Excel pour l’intégralité.",
                style_note,
            )
        )

    def pied(toile: Any, _doc: Any) -> None:
        toile.saveState()
        toile.setStrokeColor(FILET)
        toile.setLineWidth(0.4)
        toile.line(14 * mm, 11 * mm, landscape(A4)[0] - 14 * mm, 11 * mm)
        toile.setFont("Helvetica-Bold", 7)
        toile.setFillColor(MARINE)
        toile.drawString(14 * mm, 7 * mm, "CADERAC")
        toile.setFont("Helvetica", 7)
        toile.setFillColor(ENCRE_DOUX)
        toile.drawString(31 * mm, 7 * mm, definition.libelle)
        toile.drawRightString(
            landscape(A4)[0] - 14 * mm, 7 * mm, f"Page {toile.getPageNumber()}"
        )
        toile.restoreState()

    document.build(histoire, onFirstPage=pied, onLaterPages=pied)
    return tampon.getvalue()
